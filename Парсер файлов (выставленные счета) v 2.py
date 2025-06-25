import os  # Для работы с файловой системой
import openpyxl  # Для работы с Excel файлами
from openpyxl import Workbook  # Для создания новых Excel файлов
import re  # Для работы с регулярными выражениями
from datetime import datetime  # Для работы с датами
import ctypes


def process_folders():
    # Создаем новую рабочую книгу Excel для результатов
    result_wb = Workbook()
    # Получаем активный лист
    result_ws = result_wb.active
    # Переименовываем лист
    result_ws.title = "Результаты"
    # Добавляем заголовки столбцов
    result_ws.append(["Название", "Количество"])

    # Словарь для хранения данных (название: количество)
    data_dict = {}
    count = 0
    # Компилируем регулярное выражение для пропуска определенных слов
    skip_pattern = re.compile(r'^(доставка|комиссия|скидка|списание|начислить|монтажные|внести)', flags=re.IGNORECASE)

    # Список целевых директорий для обработки
    target_dirs = [
        "Все клиенты",
        "Все клиенты/_РАССРОЧКА"
    ]

    # Обрабатываем каждую директорию из списка
    for dir_path in target_dirs:
        # Проверяем существование директории
        if not os.path.exists(dir_path):
            print(f"Папка не найдена: {dir_path}")
            continue  # Переходим к следующей директории

        print(f"\nОбработка папки: {dir_path}")

        # Перебираем элементы в текущей директории
        for item in os.listdir(dir_path):
            # Полный путь к элементу
            full_path = os.path.join(dir_path, item)

            # Проверяем, что это директория и в названии нет ! или +
            if os.path.isdir(full_path) and not re.match(r'^[!+."@№#$;%:^?&*()\-=_]', item):
                print(f"  Найдена подпапка: {item}")

                # Перебираем файлы в поддиректории

                for file in os.listdir(full_path):

                    # Проверяем, что файл начинается на "проект" и имеет расширение Excel
                    if file.lower().startswith("проект") and file.lower().endswith(('.xlsx',)):
                        file_path = os.path.join(full_path, file)
                        print(f"Найден файл: {file}")
                        count += 1

                        try:
                            # Загружаем рабочую книгу Excel
                            wb = openpyxl.load_workbook(file_path, data_only=True)
                            # Получаем активный лист
                            ws = wb.active
                            # Определяем максимальное количество строк
                            max_row = ws.max_row

                            # Инициализируем счетчик пустых строк
                            empty_line_counter = 0
                            # Максимально допустимое количество пустых строк подряд
                            max_empty_lines = 2

                            # Перебираем строки в листе (начиная с 1, а не с 2)
                            for row in range(1, max_row + 1):
                                # Получаем значения из столбцов B и I
                                name = ws[f'B{row}'].value
                                qty = ws[f'I{row}'].value

                                # Проверяем, являются ли оба значения пустыми
                                if not name:
                                    empty_line_counter += 1
                                else:
                                    empty_line_counter = 0  # Сбрасываем счетчик

                                # Если найдено 2+ пустых строк подряд - прерываем цикл
                                if empty_line_counter >= max_empty_lines:
                                    print(f"Файл {file}: обнаружены 2 пустые строки подряд")
                                    break

                                # Если есть название и количество - число
                                if name and isinstance(qty, (int, float)):
                                    # Проверяем, не является ли название исключенным
                                    if isinstance(name, str) and skip_pattern.search(name):
                                        continue  # Пропускаем эту строку

                                    # Добавляем данные в словарь (суммируем количества)
                                    data_dict[name] = data_dict.get(name, 0) + qty



                        except Exception as e:
                            print(f"Ошибка при обработке файла {file_path}: {e}")

    # Получаем текущую дату
    now = datetime.now()
    # Сортируем данные по алфавиту перед сохранением
    for name in sorted(data_dict.keys(), key=lambda x: str(x).lower()):
        result_ws.append([name, data_dict[name]])

    # Формируем имя результирующего файла с датой
    result_file = f"Сводка выставленых счетов {now.date()}.xlsx"
    # Сохраняем рабочую книгу
    result_wb.save(result_file)
    print(f"\nРезультаты сохранены в файл: {result_file}")


    # Уведомление после выполнения
    # Простое сообщение
    ctypes.windll.user32.MessageBoxW(0, f"Обработано количество файлов : {count} ", "Парсинг файлов завершен", 0x40)






# Точка входа в программу
if __name__ == "__main__":
    process_folders()
