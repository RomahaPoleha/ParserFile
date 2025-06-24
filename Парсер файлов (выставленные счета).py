
import os
import openpyxl
from openpyxl import Workbook
import re
from datetime import datetime

def process_folders():
    result_wb = Workbook()
    result_ws = result_wb.active
    result_ws.title = "Результаты"
    result_ws.append(["Название", "Количество"])

    data_dict = {}
    skip_pattern = re.compile(r'^(доставка|комиссия|скидка|списание|начислить|монтажные|внести)', flags=re.IGNORECASE)

    target_dirs = [
        "Все клиенты",
        "Все клиенты/_РАССРОЧКА"
    ]

    for dir_path in target_dirs:
        if not os.path.exists(dir_path):
            print(f"Папка не найдена: {dir_path}")
            continue

        print(f"\nОбработка папки: {dir_path}")

        for item in os.listdir(dir_path):
            full_path = os.path.join(dir_path, item)

            # Изменено: проверяем только первый символ в названии папки
            if os.path.isdir(full_path) and not re.match(r'^[!+."@№#$;%:^?&*()\-=_]', item):
                print(f"  Найдена подпапка: {item}")

                for file in os.listdir(full_path):
                    if file.lower().startswith("проект") and file.lower().endswith('.xlsx'):
                        file_path = os.path.join(full_path, file)
                        print(f"Найден файл: {file}")

                        try:
                            wb = openpyxl.load_workbook(file_path, data_only=True)
                            ws = wb.active
                            max_row = ws.max_row

                            empty_line_counter = 0
                            max_empty_lines = 2

                            for row in range(1, max_row + 1):
                                name = ws[f'B{row}'].value
                                qty = ws[f'I{row}'].value

                                if name is None and qty is None:
                                    empty_line_counter += 1
                                else:
                                    empty_line_counter = 0

                                if empty_line_counter >= max_empty_lines:
                                    print(f"Файл {file}: обнаружены 2 пустые строки подряд")
                                    break

                                if name and isinstance(qty, (int, float)):
                                    if isinstance(name, str) and skip_pattern.search(name.strip()):
                                        continue

                                    data_dict[name] = data_dict.get(name, 0) + qty

                        except Exception as e:
                            print(f"Ошибка при обработке файла {file_path}: {e}")

    now = datetime.now()
    for name, qty in data_dict.items():
        result_ws.append([name, qty])

    result_file = f"Сводка за {now.date()}.xlsx"
    result_wb.save(result_file)
    print(f"\nРезультаты сохранены в файл: {result_file}")

if __name__ == "__main__":
    process_folders()
